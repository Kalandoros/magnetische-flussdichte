import sys
import traceback
import tomllib
from collections import defaultdict
from collections.abc import Hashable
from dataclasses import asdict
from pathlib import Path
from functools import lru_cache
import shutil
from typing import Any, Optional
import warnings

import pandas as pd
from openpyxl import load_workbook

from src.utils import formatter, traceback_detail, mappings

pd.options.display.float_format = '{:12.3e}'.format
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.precision', 3)

TOML_FILE = "project.toml"

TEMPLATE_DIRECTORY_NAME = "src/templates"
EXCEL_EXPORT_FILE_KURZSCHLUSSKRAFT_LEITERSEILE = "Export Vorlage Kurzschlusskraft Leiterseile.xlsx"

DATA_DIRECTORY = "src/data"
FILE_NAME_LEITERSEILDATEN = "Leiterseildaten.csv"

@lru_cache(maxsize=1)
def get_app_version() -> Optional[str]:
    """
    Liest die App-Version aus der project.toml.
    """
    toml_path = Path(get_project_root(), TOML_FILE)

    try:
        # TOML laden und Version auslesen
        with open(toml_path, "rb") as f:
            data: dict[str, Any] = tomllib.load(f)
            version = data.get("project", {}).get("version")
            return version

    except FileNotFoundError as fnfe:
        sys.stderr.write(f"Warning: Configuration file not found at {toml_path}\n")
        traceback.print_exc(limit=10, file=sys.stderr, chain=True)
        return "dev-local"
    except tomllib.TOMLDecodeError as e:
        sys.stderr.write(f"Error: pyproject.toml has invalid TOML syntax: {e}\n")
        traceback.print_exc(limit=10, file=sys.stderr, chain=True)
        return "error-syntax"
    except Exception as e:
        sys.stderr.write(f"Unexpected error loading version: {e}\n")
        traceback.print_exc(limit=10, file=sys.stderr, chain=True)
        return None

@lru_cache(maxsize=1)
def get_project_root() -> Path:
    """Gibt das Projekt-Root-Verzeichnis zurück."""
    # Aufbau: .../src/utils/dataloader.py -> Root
    return Path(__file__).parent.parent.parent

def convert_value_for_excel(value: Any) -> Any:
    """
    Konvertiert einen Wert in den passenden Typ für Excel-Export.

    Args:
        value: Zu konvertierender Wert

    Returns:
        Konvertierter Wert (str, int, float oder '')
    """
    # Prüfe auf leere Werte
    # Werte fuer den Excel-Export normalisieren
    if value is None or value in ('0', 0):
        return ''

    # Wenn bereits eine Zahl, direkt zurückgeben
    if isinstance(value, (int, float)):
        return value

    # String-Konvertierung versuchen
    if isinstance(value, str):
        try:
            # Float wenn Dezimaltrennzeichen vorhanden
            if '.' in value or ',' in value:
                return float(value.replace(',', '.'))
            # Ansonsten Integer
            return int(value)
        except (ValueError, AttributeError):
            # Bei Fehler String beibehalten
            return value

    return value

@lru_cache(maxsize=5)
def _load_csv_to_df_cached(file_name: str) -> pd.DataFrame:
    """
    Interner Cache für CSV-Importe (nur lesen, keine Mutation).
    """
    # print(get_project_root(), DATA_DIRECTORY, file_name)
    # CSV-Pfad aus dem Projekt-Root bauen
    file = Path(get_project_root(), DATA_DIRECTORY, file_name)
    if file is None:
        raise FileNotFoundError(f"Datei {file_name} konnte nicht gefunden werden!")
    raw_data = pd.read_csv(file, header=0, delimiter=";", na_filter=False)
    return pd.DataFrame(raw_data)

def load_csv_to_df_with_cache(file_name: str) -> pd.DataFrame:
    """Lädt die CSV-Daten und gibt einen DataFrame zurück (mit Cache).

    Args:
        file_name: Name der CSV-Datei

    Returns:
        DataFrame mit den geladenen Daten (keine Header-Zeile)
    """
    # Kopie zurückgeben, damit der Cache nicht versehentlich mutiert wird.
    return _load_csv_to_df_cached(file_name).copy()

def convert_df_to_dict(df: pd.DataFrame) -> list[dict[Hashable, Any]]:
    """Lädt einen DataFrame und gibt ein Dictionary zurück.

    Args:
        df: DataFrame mit den Eingabedaten (vertikales Layout)

    Returns:
        Liste mit geladenen Daten
    """
    dictionary = df.to_dict(orient='records')
    return dictionary

def load_excel_to_df(file_path: Optional[str | Path] = None) -> pd.DataFrame:
    """
    Lädt eine Excel-Datei und gibt sie als DataFrame für den Import der Excel-Vorlagen zurück.
    Die Excel-Datei hat eine vertikale Struktur:
    - Spalte 0: Parameternamen
    - Spalte 1: Werte

    Args:
        file_path: Pfad zur Excel-Datei. Wenn None, wird die Standard-Eingabe-Datei verwendet.

    Returns:
        DataFrame mit den geladenen Daten (keine Header-Zeile)
    """
    if file_path is None:
        raise TypeError("Kein Dateipfad zum Laden vorhanden!")
    else:
        file_path = Path(file_path)

    try:
        # Verwende Context Manager für das automatische schliessen
        with pd.ExcelFile(file_path, engine='openpyxl') as xlsx:
            raw_data = pd.read_excel(xlsx, header=None, na_filter=False)
            df = pd.DataFrame(raw_data)
        return df

    except FileNotFoundError as fnfe:
        error_msg = traceback_detail.get_exception_message(fnfe)
        sys.stderr.write(f"{error_msg}\n")
        traceback.print_exc(limit=10, file=sys.stderr, chain=True)
        return pd.DataFrame()
    except PermissionError as pe:
        error_msg = traceback_detail.get_exception_message(pe)
        sys.stderr.write(f"{error_msg}\n")
        traceback.print_exc(limit=10, file=sys.stderr, chain=True)
        return pd.DataFrame()
    except ValueError as ve:
        error_msg = traceback_detail.get_exception_message(ve)
        sys.stderr.write(f"{error_msg}\n")
        traceback.print_exc(limit=10, file=sys.stderr, chain=True)
        return pd.DataFrame()
    except Exception as e:
        error_msg = traceback_detail.get_exception_message(e)
        sys.stderr.write(f"{error_msg}\n")
        traceback.print_exc(limit=10, file=sys.stderr, chain=True)
        return pd.DataFrame()

def convert_excel_to_dict_with_mapping(df: pd.DataFrame, mapping: str) -> tuple[dict[str, Any], list[str], list[str]]:
    """
    Konvertierung des DataFrame aus der Excel-Vorlage in ein Dictionary,
    das direkt zum Setzen der Taipy GUI Widgets verwendet werden kann.

    Die Excel-Datei hat eine vertikale Struktur:
    - Spalte 0: Parameternamen (z.B. "Art der Leiterseilbefestigung")
    - Spalte 1: Werte (z.B. "Abgespannt")

    Args:
        df: DataFrame mit den Eingabedaten (vertikales Layout)
        mapping: Name des Excel-Mappings, das in mapping.json verwendet werden soll

    Returns:
        Tuple mit (input_dict, loaded_fields, skipped_fields)
        - input_dict: Dictionary mit den Eingabewerten
        - loaded_fields: Liste der erfolgreich geladenen Felder
        - skipped_fields: Liste der übersprungenen Felder (leer oder nicht gefunden)
    """
    if df.empty:
        return {}, [], []

    # Erstelle ein Dictionary aus den Zeilen: Spalte 0 = Key, Spalte 1 = Value
    excel_data: dict[str, Any] = {}
    for idx, row in df.iterrows():
        if pd.notna(row[0]) and row[0] != '':
            key = str(row[0]).strip()
            value = row[1] if len(row) > 1 and pd.notna(row[1]) and row[1] != '' else None
            if value is not None:
                excel_data[key] = value

    # Erstelle ein Dictionary mit allen relevanten Feldern
    input_dict: dict[str, Any] = {}
    loaded_fields: list[str] = []
    skipped_fields: list[str] = []

    # Mapping der Excel-Zeilennamen zu den State-Variablennamen
    field_mapping = mappings.get_mapping(mapping)

    # Konvertiere alle vorhandenen Felder
    for state_var, excel_label in field_mapping.items():
        if excel_label in excel_data:
            value = excel_data[excel_label]
            input_dict[state_var] = value
            loaded_fields.append(excel_label)
        else:
            skipped_fields.append(excel_label)

    return input_dict, loaded_fields, skipped_fields

def export_dict_to_excel_with_reversemapping(input_dict: dict[str, Any], template_path: str | Path, output_path: str | Path) -> bool:
    """
    Exportiert Dictionary in Excel-Datei basierend auf Vorlage.

    Args:
        input_dict: Dictionary mit State-Variablennamen als Keys und Werten
        template_path: Pfad zur Vorlagen-Excel-Datei, liefert auch den Key fürs reverse_mapping.json an
        output_path: Pfad zur Ausgabe-Excel-Datei

    Returns:
        True bei Erfolg, False bei Fehler
    """
    warnings.filterwarnings("ignore", category=ResourceWarning, message=".*unclosed file.*")

    template_path = Path(template_path)
    if not template_path.exists():
        print(f"Vorlage nicht gefunden: {template_path}")
        return False

    # Gibt das dict aus reverse_mapping.json zurück
    reverse_field_mapping = mappings.get_reverse_mapping(template_path)

    # Kopiere die Vorlage zur Ausgabedatei (um Formatierung zu erhalten)
    shutil.copy(template_path, output_path)

    # Öffne die kopierte Datei mit openpyxl (behält Formatierung)
    wb = None
    try:
        wb = load_workbook(output_path, keep_vba=False, data_only=False)
        ws = wb.active

        # Durchlaufe alle Zeilen in der Excel-Datei
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            # Spalte A (Index 0) enthält die Labels
            if row[0].value:
                excel_label = str(row[0].value).strip()

                # Suche nach dem passenden State-Variablennamen
                for mapped_label, state_var in reverse_field_mapping.items():
                    if excel_label == mapped_label:
                        # Hole den Wert aus den Werten von reverse_field_mapping dict auf Basis von mappingreversed.json
                        value = input_dict.get(state_var, '')

                        # Konvertiere den Wert in den richtigen Typ
                        if value is None or value == '0' or value == "0" or value == 0:
                            value = ''
                        elif isinstance(value, str):
                            # Versuche String in Zahl zu konvertieren, wenn möglich
                            try:
                                # Versuche zuerst Float
                                if '.' in value or ',' in value:
                                    value = float(value.replace(',', '.'))
                                else:
                                    # Versuche Int zu konvertieren
                                    value = int(value)
                            except (ValueError, AttributeError):
                                # Wenn nicht konvertierbar, behalte String
                                pass

                        # Setze den Wert in Spalte B (Index 1)
                        row[1].value = value
                        break

        # Speichere die Datei (Formatierung bleibt erhalten)
        wb.save(output_path)
        return True
    except Exception as e:
        error_msg = traceback_detail.get_exception_message(e)
        sys.stderr.write(f"{error_msg}\n")
        print(f"Fehler beim Exportieren der Excel-Datei: {e}")
        traceback.print_exc(limit=10, file=sys.stderr, chain=True)
        return False
    finally:
        # Stell sicher, dass die Datei geschlossen wird, auch bei einem Fehler
        if wb is not None:
            wb.close()

def create_df_from_calc_results_kurzschlusskreafte_leiterseile(calc_result: dict[str, Any], temp_low: str, temp_high: str) -> pd.DataFrame:
    """
    Erstellt einen formatierten DataFrame aus den Berechnungsergebnissen.

    Args:
        calc_result: Dictionary mit 'F_st_20' und 'F_st_80' als Keys
        temp_low: String of the selected temperature
        temp_high: String of the selected temperature
    Returns:
        DataFrame mit formatierten Ergebnissen
    """
    temp_low_with_unit = str(temp_low) + " " + "°C"
    temp_high_with_unit = str(temp_high) + " " + "°C"

    # Hole beide Ergebnisse
    result_20 = calc_result.get('F_st_20')
    result_80 = calc_result.get('F_st_80')

    if not result_20 or not result_80:
        return pd.DataFrame({'Hinweis': ['Keine Berechnungsergebnisse verfügbar']})

    # Konvertiere zu Dictionaries
    dict_20 = asdict(result_20)
    dict_80 = asdict(result_80)

    # Erstelle Liste für DataFrame-Rows
    data = []

    # Iteriere durch alle Felder
    for param_name in dict_20.keys():
        val_20 = dict_20[param_name]
        val_80 = dict_80[param_name]

        # Nur nicht-None Werte hinzufügen
        if val_20 is not None and val_80 is not None:
            # WICHTIG: Konvertiere Sympy-Objekte zu nativen Python-Typen
            # Sympy.Float ist nicht JSON-serialisierbar für Taipy
            if hasattr(val_20, '__float__'):
                val_20 = float(val_20)
            if hasattr(val_80, '__float__'):
                val_80 = float(val_80)

            # Formatiere mit der Formatter-Funktion
            val_20_formatted = formatter.format_number_nice_to_string_for_repr(val_20)
            val_80_formatted = formatter.format_number_nice_to_string_for_repr(val_80)

            data.append({
                'Parameter': param_name,
                temp_low_with_unit: val_20_formatted,
                temp_high_with_unit: val_80_formatted
                #'-20°C': val_20,
                #'80°C': val_80
            })
    # Erstelle DataFrame
    df = pd.DataFrame(data)

    # Entfernen der Einrückung dieses Teils und des darüber liegenden Teils, wenn ein numerischer Export erforderlich ist
    """
        if not df.empty:
            df['Parameter'] = df['Parameter'].astype(str)
            df['-20°C'] = pd.to_numeric(df['-20°C'], errors='coerce')
            df['80°C'] = pd.to_numeric(df['80°C'], errors='coerce')
    """

    return df


if __name__ == "__main__":
    pass
    # Liefert alle Werte einer Spalte zurück
    # print(load_csv_to_df_with_cache()["Bezeichnung"])

    # Liefert die Bezeichnung eines speziellen Indexes zurück
    # print(load_csv_to_df_with_cache().at[2, "Bezeichnung"])

    # Liefert den Wert einer Spalte basierend auf dem Index zurück
    # print(load_csv_to_df_with_cache().iloc[2]["Querschnitt eines Teilleiters"])

    # print(load_csv_to_df_with_cache())

    # print(convert_df_to_dict(load_csv_to_df_with_cache()))
    # list_dict_leiterseile = convert_df_to_dict(load_csv_to_df_with_cache())
    # print(list_dict_leiterseile[0])
    # dict_leiterseile = list_dict_leiterseile[0]

    # print(list(dict_leiterseile.keys()))
    # print(list(dict_leiterseile.values()))
    # print(dict_leiterseile.get("Bezeichnung"))
    # print(dict_leiterseile["Bezeichnung"])
